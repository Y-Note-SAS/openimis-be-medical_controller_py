from .apps import MedicalControllerConfig
# from django.core.exceptions import PermissionDenied
from django.utils.translation import gettext as _
from core.security import checkUserWithRights
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework.decorators import (
    api_view,
    permission_classes
)
from datetime import datetime
from .models import (
    MedicalControlMission,
    MissionHealthFacility,
    FilteredClaimsForMission
)

@api_view(["GET"])
@permission_classes(
    [
        checkUserWithRights(
            MedicalControllerConfig.gql_mutation_medical_controller_perms,
        )
    ]
)
def download_mission(request, mission_code):
    mission = (
        MedicalControlMission.objects
        .select_related("user")
        .get(mission_code=mission_code)
    )
    print("ABC")

    wb = Workbook()
    ws = wb.active
    ws.title = _("Mission")

    bold = Font(bold=True)

    # =====================================================
    # EN-TETE
    # =====================================================

    row = 1

    ws.cell(row=row, column=1, value=_("Nom de la mission"))
    ws.cell(row=row, column=2, value=mission.mission_code)

    ws.cell(row=row, column=8, value=_("Téléchargé le"))
    ws.cell(row=row, column=9, value = datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

    row += 1

    ws.cell(row=row, column=1, value=_("Utilisateur"))
    ws.cell(
        row=row,
        column=2,
        value=str(mission.user.username) if mission.user else "",
    )

    row += 1

    ws.cell(row=row, column=1, value=_("Pourcentage 1"))
    ws.cell(row=row, column=2, value=mission.percentage_one)

    row += 1

    ws.cell(row=row, column=1, value=_("Pourcentage 2"))
    ws.cell(row=row, column=2, value=mission.percentage_two)

    row += 1

    ws.cell(row=row, column=1, value=_("Pourcentage 3"))
    ws.cell(row=row, column=2, value=mission.percentage_three)

    row += 1

    ws.cell(row=row, column=1, value=_("Pourcentage 4"))
    ws.cell(row=row, column=2, value=mission.percentage_four)

    row += 2

    facilities = MissionHealthFacility.objects.filter(
        mission=mission
    ).select_related("health_facility")

    ws.cell(row=row, column=1, value=_("FOSA concernées"))
    ws.cell(row=row, column=1).font = bold

    row += 1
    health_facilities = []

    for mission_hf in facilities:
        health_facilities.append(mission_hf.health_facility.id)
        ws.cell(
            row=row,
            column=1,
            value=str(mission_hf.health_facility)
        )
        row += 1

    row += 3

    # =====================================================
    # LISTE DES CLAIMS
    # =====================================================

    headers = [
        _("Numéro du claim"),
        _("Nom de la FOSA"),
        _("Numéro de l'assuré"),
        _("Statut du claim"),
        _("Audité"),
        _("Montant réclamé"),
        _("Montant approuvé"),
        _("Montant audité"),
        _("Catégorie"),
        _("Statut de l'audit"),
        _("Motif de rejet"),
        _("Raison de rejet"),
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = bold

    row += 1

    claims_for_mission = FilteredClaimsForMission.objects.filter(
        claim__health_facility__id__in=health_facilities,
        mission=mission
    )

    for claim in claims_for_mission:
        insuree_number = ""

        if claim.claim.insuree:
            insuree_number = claim.claim.insuree.chf_id

        ws.cell(row=row, column=1, value=claim.claim.code)
        ws.cell(
            row=row,
            column=2,
            value=str(claim.claim.health_facility)
            if claim.claim.health_facility
            else "",
        )
        ws.cell(row=row, column=3, value=insuree_number)
        ws.cell(row=row, column=4, value=claim.claim.status)
        ws.cell(
            row=row,
            column=5,
            value=_("Oui") if claim.claim.audited else _("Non"),
        )
        ws.cell(row=row, column=6, value=claim.claim.claimed)
        ws.cell(row=row, column=7, value=claim.claim.approved)
        ws.cell(row=row, column=8, value=claim.claim.amount_audited)
        ws.cell(row=row, column=9, value=claim.claim.claim_category)
        ws.cell(row=row, column=10, value=claim.claim.audit_status)
        ws.cell(row=row, column=11, value=claim.claim.rejection_motive)
        ws.cell(
            row=row,
            column=12,
            value=claim.claim.rejection_reason_after_audit,
        )

        row += 1

    # Ajustement largeur colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = max_length + 5

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="mission_{mission.mission_code}.xlsx"'
    )

    wb.save(response)

    return response
